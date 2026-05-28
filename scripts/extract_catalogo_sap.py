"""
Extrae la pestaña "codigos sap y material" del Excel de la BD de compras
y genera catalogo_sap.json en public/data/.

Uso:
  python scripts/extract_catalogo_sap.py

Asegúrate de que el Excel actualizado está en:
  C:\\Users\\juank\\BASE_DATOS_COMPRAS_PROVEEDORES_v2.xlsx
  (o actualiza EXCEL_PATH si lo tienes en otro sitio)
"""

import openpyxl, json, sys, os
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH  = r"C:\Users\juank\BASE_DATOS_COMPRAS_PROVEEDORES_v2.xlsx"
OUT_DIR     = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
OUT_FILE    = os.path.join(OUT_DIR, 'catalogo_sap.json')

# Nombres posibles de la pestaña (en caso de ligeras variaciones)
SHEET_NAMES = [
    'codigos sap y material',
    'Codigos SAP y Material',
    'CODIGOS SAP Y MATERIAL',
    'Códigos SAP y Material',
    'Códigos SAP',
    'SAP Codigos',
    'codigos_sap',
]

# Columnas posibles para el código SAP
CODIGO_KEYS = [
    'Código SAP', 'Codigo SAP', 'CÓDIGO SAP', 'codigo sap',
    'Código', 'Codigo', 'Material', 'Nº Material', 'Número Material',
    'Número de material', 'Cod. SAP', 'SAP', 'Código de material',
]

# Columnas posibles para la descripción
DESC_KEYS = [
    'Descripción Material', 'Descripcion Material', 'DESCRIPCIÓN MATERIAL',
    'Descripción', 'Descripcion', 'DESCRIPCIÓN', 'Texto breve de material',
    'Texto breve material', 'Texto breve', 'Descripcion del material',
    'Desc. Material', 'Nombre', 'Denominación',
]


def find_col(headers: list, candidates: list) -> str | None:
    """Devuelve el primer nombre de columna que coincida (insensible a mayúsculas/tildes)."""
    def norm(s: str) -> str:
        import unicodedata
        s = unicodedata.normalize('NFD', s.lower())
        return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

    h_norm = {norm(h): h for h in headers}
    for c in candidates:
        cn = norm(c)
        if cn in h_norm:
            return h_norm[cn]
    return None


def load_sheet(wb: openpyxl.Workbook) -> list[dict]:
    """Encuentra y carga la pestaña del catálogo."""
    import unicodedata
    def norm(s): return unicodedata.normalize('NFD', s.lower()).encode('ascii', 'ignore').decode()

    # Buscar la pestaña
    sheet = None
    for sn in wb.sheetnames:
        for candidate in SHEET_NAMES:
            if norm(sn) == norm(candidate):
                sheet = wb[sn]
                print(f"✓ Pestaña encontrada: '{sn}'")
                break
        if sheet:
            break

    if sheet is None:
        print(f"\nPestañas disponibles en el Excel:")
        for sn in wb.sheetnames:
            rows = sheet = wb[sn]
            print(f"  - '{sn}'")
        raise ValueError(
            f"No se encontró la pestaña del catálogo SAP.\n"
            f"Nombres buscados: {SHEET_NAMES}\n"
            f"Pestañas disponibles: {wb.sheetnames}"
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Primera fila = cabeceras
    headers = [str(c).strip() if c is not None else '' for c in rows[0]]
    print(f"  Columnas detectadas: {[h for h in headers if h]}")

    ck = find_col(headers, CODIGO_KEYS)
    dk = find_col(headers, DESC_KEYS)

    if not ck:
        print(f"\n⚠ No se encontró columna de código SAP. Columnas: {headers}")
        print("  Usando columna 0 como código y columna 1 como descripción")
        ci, di = 0, 1
    else:
        ci = headers.index(ck)
        di = headers.index(dk) if dk else (1 if ci == 0 else 0)
        print(f"  Columna código: '{ck}' (col {ci})")
        print(f"  Columna descripción: '{dk}' (col {di})")

    result = []
    for row in rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        codigo = str(row[ci]).strip() if ci < len(row) and row[ci] is not None else ''
        desc   = str(row[di]).strip() if di < len(row) and row[di] is not None else ''
        if not codigo or codigo in ('None', '', 'nan'):
            continue
        if not desc or desc in ('None', '', 'nan'):
            continue
        # Solo SAP numéricos (para filtrar cabeceras duplicadas u otros datos)
        if not codigo.replace(' ', '').isdigit():
            continue
        result.append({'codigo': codigo.strip(), 'descripcion': desc.strip()})

    return result


def main():
    print(f"Leyendo Excel: {EXCEL_PATH}")
    if not os.path.exists(EXCEL_PATH):
        print(f"\n⚠ No se encuentra el Excel en:\n  {EXCEL_PATH}")
        print("  Actualiza EXCEL_PATH en este script.")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    registros = load_sheet(wb)

    if not registros:
        print("⚠ No se extrajeron registros. Revisa la pestaña y las columnas.")
        sys.exit(1)

    # Deduplicar por código (mantener primera aparición)
    vistos = set()
    dedup = []
    for r in registros:
        if r['codigo'] not in vistos:
            vistos.add(r['codigo'])
            dedup.append(r)

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(dedup, f, ensure_ascii=False, indent=2)

    print(f"\n✓ {len(dedup)} códigos SAP exportados → {OUT_FILE}")

    # Muestra sample
    print("\nPrimeros 5 registros:")
    for r in dedup[:5]:
        print(f"  {r['codigo']:>12}  {r['descripcion'][:60]}")


if __name__ == '__main__':
    main()
